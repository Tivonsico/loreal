from __future__ import annotations

import hashlib
import io
from datetime import UTC, date, datetime, timedelta, timezone
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.backend.models import (
    AdverseReactionDetail,
    Conversation,
    ImportBatch,
    ImportRowError,
    LogisticsDetail,
    Message,
    OfflinePaymentDetail,
    Order,
    Product,
    ReplacementDetail,
    ReturnDetail,
    WorkOrder,
    utc_now,
)

SHANGHAI = timezone(timedelta(hours=8), name="Asia/Shanghai")
MAX_WORKBOOK_BYTES = 20 * 1024 * 1024
SHEET_HEADERS = {
    "聊天记录": [
        "会话ID",
        "消息序号",
        "message_id",
        "发送时间",
        "角色",
        "买家昵称",
        "发送方",
        "店铺",
        "scene_major",
        "scene_minor",
        "is_target_buyer_message",
        "message_text",
        "内容类型",
        "chat_content",
        "category",
        "image_path",
        "关联订单号",
        "关联工单号",
    ],
    "订单": [
        "订单号",
        "会话ID",
        "买家昵称",
        "店铺",
        "商品货号",
        "商品名称",
        "数量",
        "单价(元)",
        "实付金额(元)",
        "订单状态",
        "下单时间",
        "付款时间",
        "发货时间",
        "快递公司",
        "物流单号",
        "收货省",
        "收货市",
        "赠品",
        "买家留言",
    ],
    "补发换货工单": [
        "工单号",
        "会话ID",
        "关联订单号",
        "买家昵称",
        "店铺",
        "工单类型",
        "售后原因",
        "发出商品货号",
        "发出商品名称",
        "数量",
        "原订单物流单号",
        "补发物流单号",
        "快递公司",
        "发货仓库",
        "客诉加急",
        "工单状态",
        "处理人",
        "创建时间",
        "完成时间",
    ],
    "线下打款工单": [
        "工单号",
        "会话ID",
        "关联订单号",
        "买家昵称",
        "店铺",
        "打款类型",
        "退款问题类型",
        "退款金额(元)",
        "支付宝实名",
        "支付宝账号",
        "相关物流单号",
        "转账状态",
        "工单状态",
        "处理人",
        "创建时间",
        "完成时间",
    ],
    "物流工单": [
        "工单号",
        "会话ID",
        "关联订单号",
        "买家昵称",
        "店铺",
        "问题类型",
        "快递公司",
        "问题包裹物流单号",
        "发货仓",
        "订单实付(元)",
        "处理方案",
        "收货省",
        "收货市",
        "工单状态",
        "处理人",
        "创建时间",
        "完成时间",
    ],
    "不良反应工单": [
        "工单号",
        "会话ID",
        "关联订单号",
        "买家昵称",
        "店铺",
        "类型",
        "年龄",
        "肤质",
        "使用商品",
        "产品批次号",
        "不适部位",
        "症状描述",
        "用后多久出现",
        "是否停用",
        "是否就医",
        "任务状态",
        "处理人",
        "创建时间",
        "完成时间",
    ],
    "售后退货工单": [
        "工单号",
        "会话ID",
        "关联订单号",
        "买家昵称",
        "店铺",
        "包裹类型",
        "退货原因",
        "退货物流单号",
        "快递公司",
        "退款编号",
        "签收建议",
        "是否异常",
        "任务状态",
        "处理人",
        "创建时间",
        "完成时间",
    ],
}
WORK_ORDER_TYPES = {
    "补发换货工单": "replacement_exchange",
    "线下打款工单": "offline_payment",
    "物流工单": "logistics",
    "不良反应工单": "adverse_reaction",
    "售后退货工单": "after_sale_return",
}
ROLE_MAP = {"买家": "customer", "客服": "customer_service", "系统推送": "system"}


def _text(value: Any) -> str | None:
    if value is None:
        return None
    result = str(value).strip()
    return result or None


def _integer(value: Any) -> int | None:
    if value in (None, ""):
        return None
    return int(value)


def _decimal(value: Any) -> str | None:
    if value in (None, ""):
        return None
    try:
        return str(Decimal(str(value)).quantize(Decimal("0.01")))
    except InvalidOperation as exc:
        raise ValueError(f"不是有效金额：{value}") from exc


def _boolean(value: Any) -> bool | None:
    text = _text(value)
    if text is None:
        return None
    if text in {"是", "是的", "true", "True", "1", "加急"}:
        return True
    if text in {"否", "false", "False", "0", "不加急"}:
        return False
    raise ValueError(f"不是有效的是/否值：{value}")


def _datetime(value: Any) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        parsed = value
    elif isinstance(value, date):
        parsed = datetime.combine(value, datetime.min.time())
    else:
        raw = str(value).strip()
        for suffix in ("（定金）", "(定金)"):
            raw = raw.removesuffix(suffix).strip()
        parsed = None
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d"):
            try:
                parsed = datetime.strptime(raw, fmt)
                break
            except ValueError:
                continue
        if parsed is None:
            raise ValueError(f"不是支持的日期时间：{value}")
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=SHANGHAI)
    return parsed.astimezone(UTC).isoformat()


def _status(source_status: Any) -> str:
    value = _text(source_status) or ""
    if "完结" in value or "成功" in value:
        return "completed"
    if "待" in value:
        return "pending"
    return "processing"


def _rows(worksheet) -> list[tuple[int, dict[str, Any]]]:  # noqa: ANN001
    values = worksheet.iter_rows(values_only=True)
    headers = [_text(value) for value in next(values)]
    result = []
    for row_number, row in enumerate(values, start=2):
        if not any(value not in (None, "") for value in row):
            continue
        result.append((row_number, dict(zip(headers, row, strict=False))))
    return result


def _error(
    errors: list[dict[str, Any]],
    sheet: str,
    row: int,
    code: str,
    message: str,
    column: str | None = None,
) -> None:
    errors.append(
        {
            "sheet_name": sheet,
            "row_number": row,
            "column_name": column,
            "error_code": code,
            "message": message,
        }
    )


def _parse_chat(row: dict[str, Any]) -> dict[str, Any]:
    role_source = _text(row["角色"])
    if role_source not in ROLE_MAP:
        raise ValueError(f"未知角色：{role_source}")
    content_type = _text(row["内容类型"])
    return {
        "conversation_source_id": _text(row["会话ID"]),
        "sequence_no": _integer(row["消息序号"]),
        "source_external_id": _text(row["message_id"]),
        "created_at": _datetime(row["发送时间"]),
        "sender_role": ROLE_MAP[role_source],
        "buyer_nickname": _text(row["买家昵称"]),
        "message_type": "image" if content_type == "图片" else "text",
        "content": _text(row["message_text"]),
        "raw_content": _text(row["chat_content"]),
        "media_url": _text(row["image_path"]) if content_type == "图片" else None,
        "related_order_external_id": _text(row["关联订单号"]),
        "related_work_order_external_id": _text(row["关联工单号"]),
        "source": {
            "sender": _text(row["发送方"]),
            "shop": _text(row["店铺"]),
            "scene_major": _text(row["scene_major"]),
            "scene_minor": _text(row["scene_minor"]),
            "category": _text(row["category"]),
        },
    }


def _parse_order(row: dict[str, Any]) -> dict[str, Any]:
    paid_source = _text(row["付款时间"])
    return {
        "external_id": _text(row["订单号"]),
        "conversation_source_id": _text(row["会话ID"]),
        "buyer_nickname": _text(row["买家昵称"]),
        "shop": _text(row["店铺"]),
        "product_external_id": _text(row["商品货号"]),
        "product_name": _text(row["商品名称"]),
        "quantity": _integer(row["数量"]) or 1,
        "unit_price": _decimal(row["单价(元)"]),
        "total_amount": _decimal(row["实付金额(元)"]),
        "status": _text(row["订单状态"]) or "unknown",
        "ordered_at": _datetime(row["下单时间"]),
        "paid_at": _datetime(row["付款时间"]),
        "payment_stage": "deposit" if paid_source and "定金" in paid_source else None,
        "shipped_at": _datetime(row["发货时间"]),
        "logistics_company": _text(row["快递公司"]),
        "logistics_no": _text(row["物流单号"]),
        "extra": {
            "province": _text(row["收货省"]),
            "city": _text(row["收货市"]),
            "gift": _text(row["赠品"]),
            "buyer_note": _text(row["买家留言"]),
            "shop": _text(row["店铺"]),
        },
    }


def _parse_work_order(sheet: str, row: dict[str, Any]) -> dict[str, Any]:
    status_column = "任务状态" if sheet in {"不良反应工单", "售后退货工单"} else "工单状态"
    common = {
        "external_id": _text(row["工单号"]),
        "ticket_type": WORK_ORDER_TYPES[sheet],
        "conversation_source_id": _text(row["会话ID"]),
        "order_external_id": _text(row["关联订单号"]),
        "buyer_nickname": _text(row["买家昵称"]),
        "source_status": _text(row[status_column]),
        "status": _status(row[status_column]),
        "assignee": _text(row["处理人"]),
        "opened_at": _datetime(row["创建时间"]),
        "closed_at": _datetime(row["完成时间"]),
        "source_extra": {"shop": _text(row["店铺"])},
    }
    if sheet == "补发换货工单":
        common["description"] = _text(row["售后原因"])
        common["detail"] = {
            "issue_kind": _text(row["工单类型"]),
            "product_external_id": _text(row["发出商品货号"]),
            "product_name": _text(row["发出商品名称"]),
            "quantity": _integer(row["数量"]),
            "original_tracking_no": _text(row["原订单物流单号"]),
            "replacement_tracking_no": _text(row["补发物流单号"]),
            "logistics_company": _text(row["快递公司"]),
            "warehouse": _text(row["发货仓库"]),
            "is_urgent": bool(_boolean(row["客诉加急"])),
        }
    elif sheet == "线下打款工单":
        common["description"] = _text(row["退款问题类型"])
        common["detail"] = {
            "payment_type": _text(row["打款类型"]),
            "reason": _text(row["退款问题类型"]),
            "amount": _decimal(row["退款金额(元)"]),
            "masked_real_name": _text(row["支付宝实名"]),
            "masked_account": _text(row["支付宝账号"]),
            "related_tracking_no": _text(row["相关物流单号"]),
            "transfer_status": _text(row["转账状态"]),
        }
    elif sheet == "物流工单":
        common["description"] = _text(row["问题类型"])
        common["detail"] = {
            "issue_kind": _text(row["问题类型"]),
            "logistics_company": _text(row["快递公司"]),
            "tracking_no": _text(row["问题包裹物流单号"]),
            "warehouse": _text(row["发货仓"]),
            "order_amount": _decimal(row["订单实付(元)"]),
            "handling_plan": _text(row["处理方案"]),
            "province": _text(row["收货省"]),
            "city": _text(row["收货市"]),
        }
    elif sheet == "不良反应工单":
        common["description"] = _text(row["症状描述"])
        common["detail"] = {
            "channel": _text(row["类型"]),
            "age": _integer(row["年龄"]),
            "skin_type": _text(row["肤质"]),
            "product_name": _text(row["使用商品"]),
            "product_batch_no": _text(row["产品批次号"]),
            "affected_area": _text(row["不适部位"]),
            "symptoms": _text(row["症状描述"]),
            "onset_after": _text(row["用后多久出现"]),
            "stopped_use": _text(row["是否停用"]),
            "sought_medical_care": _boolean(row["是否就医"]),
        }
    else:
        common["description"] = _text(row["退货原因"])
        common["detail"] = {
            "package_type": _text(row["包裹类型"]),
            "reason": _text(row["退货原因"]),
            "return_tracking_no": _text(row["退货物流单号"]),
            "logistics_company": _text(row["快递公司"]),
            "refund_external_id": _text(row["退款编号"]),
            "receipt_advice": _text(row["签收建议"]),
            "is_abnormal": bool(_boolean(row["是否异常"])),
        }
    return common


def parse_workbook(content: bytes) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    errors: list[dict[str, Any]] = []
    payload: dict[str, Any] = {"messages": [], "orders": [], "work_orders": []}
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        return payload, [
            {
                "sheet_name": "工作簿",
                "row_number": 0,
                "column_name": None,
                "error_code": "invalid_workbook",
                "message": f"无法读取工作簿：{exc}",
            }
        ]

    for sheet, expected in SHEET_HEADERS.items():
        if sheet not in workbook.sheetnames:
            _error(errors, sheet, 1, "missing_sheet", f"缺少工作表：{sheet}")
            continue
        actual = [_text(cell.value) for cell in next(workbook[sheet].iter_rows())]
        missing = [header for header in expected if header not in actual]
        if missing:
            _error(errors, sheet, 1, "missing_headers", f"缺少表头：{', '.join(missing)}")
            continue
        for row_number, row in _rows(workbook[sheet]):
            try:
                if sheet == "聊天记录":
                    payload["messages"].append(_parse_chat(row))
                elif sheet == "订单":
                    payload["orders"].append(_parse_order(row))
                else:
                    item = _parse_work_order(sheet, row)
                    item["source_sheet"] = sheet
                    payload["work_orders"].append(item)
            except (TypeError, ValueError, KeyError) as exc:
                _error(errors, sheet, row_number, "invalid_value", str(exc))

    _validate_relations(payload, errors)
    return payload, errors


def _validate_relations(payload: dict[str, Any], errors: list[dict[str, Any]]) -> None:
    def unique(items: list[dict[str, Any]], key: str, sheet: str) -> dict[str, dict[str, Any]]:
        found: dict[str, dict[str, Any]] = {}
        for index, item in enumerate(items, start=2):
            value = item.get(key)
            if not value:
                _error(errors, sheet, index, "required", f"{key} 不能为空", key)
            elif value in found:
                _error(errors, sheet, index, "duplicate", f"重复标识：{value}", key)
            else:
                found[value] = item
        return found

    messages = unique(payload["messages"], "source_external_id", "聊天记录")
    orders = unique(payload["orders"], "external_id", "订单")
    work_orders = unique(payload["work_orders"], "external_id", "工单")
    conversation_nicknames: dict[str, str | None] = {}
    sequences: set[tuple[str | None, int | None]] = set()
    for index, message in enumerate(messages.values(), start=2):
        conversation_id = message["conversation_source_id"]
        pair = (conversation_id, message["sequence_no"])
        if pair in sequences:
            _error(errors, "聊天记录", index, "duplicate_sequence", f"会话消息序号重复：{pair}")
        sequences.add(pair)
        nickname = message["buyer_nickname"]
        previous = conversation_nicknames.setdefault(conversation_id, nickname)
        if previous != nickname:
            _error(
                errors,
                "聊天记录",
                index,
                "nickname_mismatch",
                f"会话 {conversation_id} 的买家昵称不一致",
            )
        if (
            message["related_order_external_id"]
            and message["related_order_external_id"] not in orders
        ):
            _error(errors, "聊天记录", index, "missing_order", "关联订单号不存在")
        if (
            message["related_work_order_external_id"]
            and message["related_work_order_external_id"] not in work_orders
        ):
            _error(errors, "聊天记录", index, "missing_work_order", "关联工单号不存在")

    order_conversations: set[str] = set()
    for index, order in enumerate(orders.values(), start=2):
        conversation_id = order["conversation_source_id"]
        if conversation_id not in conversation_nicknames:
            _error(errors, "订单", index, "missing_conversation", f"会话不存在：{conversation_id}")
        if conversation_id in order_conversations:
            _error(errors, "订单", index, "multiple_orders", f"会话存在多笔订单：{conversation_id}")
        order_conversations.add(conversation_id)
        if conversation_nicknames.get(conversation_id) != order["buyer_nickname"]:
            _error(
                errors,
                "订单",
                index,
                "nickname_mismatch",
                f"会话 {conversation_id} 的买家昵称不一致",
            )

    work_order_conversations: set[str] = set()
    for index, item in enumerate(work_orders.values(), start=2):
        conversation_id = item["conversation_source_id"]
        if conversation_id not in conversation_nicknames:
            _error(
                errors,
                item["source_sheet"],
                index,
                "missing_conversation",
                f"会话不存在：{conversation_id}",
            )
        if conversation_id in work_order_conversations:
            _error(
                errors,
                item["source_sheet"],
                index,
                "multiple_work_orders",
                f"会话存在多张工单：{conversation_id}",
            )
        work_order_conversations.add(conversation_id)
        if item["order_external_id"] not in orders:
            _error(
                errors,
                item["source_sheet"],
                index,
                "missing_order",
                f"订单不存在：{item['order_external_id']}",
            )
        if conversation_nicknames.get(conversation_id) != item["buyer_nickname"]:
            _error(
                errors,
                item["source_sheet"],
                index,
                "nickname_mismatch",
                f"会话 {conversation_id} 的买家昵称不一致",
            )


def create_preview(db: Session, filename: str, content: bytes) -> ImportBatch:
    payload, errors = parse_workbook(content)
    public_summary = {
        "sheets": {"聊天记录": len(payload["messages"]), "订单": len(payload["orders"])},
        "work_order_types": {},
        "error_count": len(errors),
        "errors": errors[:100],
    }
    for work_order_type in WORK_ORDER_TYPES.values():
        public_summary["work_order_types"][work_order_type] = sum(
            item["ticket_type"] == work_order_type for item in payload["work_orders"]
        )
    batch = ImportBatch(
        filename=filename,
        file_sha256=hashlib.sha256(content).hexdigest(),
        status="ready" if not errors else "invalid",
        summary={"public": public_summary, "_payload": payload},
    )
    db.add(batch)
    db.flush()
    for item in errors:
        db.add(ImportRowError(batch_id=batch.id, **item))
    db.commit()
    db.refresh(batch)
    return batch


def _as_datetime(value: str | None) -> datetime | None:
    return datetime.fromisoformat(value) if value else None


def _upsert_detail(db: Session, model, work_order_id: int, values: dict[str, Any]) -> None:  # noqa: ANN001
    record = db.get(model, work_order_id)
    if record is None:
        db.add(model(work_order_id=work_order_id, **values))
    else:
        for key, value in values.items():
            setattr(record, key, value)


def commit_preview(db: Session, batch: ImportBatch) -> dict[str, Any]:
    if batch.status == "committed":
        return batch.summary["commit_result"]
    if batch.status != "ready":
        raise ValueError("该预览含错误，不能提交")
    payload = batch.summary["_payload"]
    counts = {"conversations": 0, "messages": 0, "orders": 0, "work_orders": 0, "products": 0}

    conversations: dict[str, Conversation] = {}
    for item in payload["messages"]:
        source_id = item["conversation_source_id"]
        if source_id in conversations:
            continue
        record = db.scalar(select(Conversation).where(Conversation.source_external_id == source_id))
        if record is None:
            record = Conversation(
                source_external_id=source_id,
                customer_id=item["buyer_nickname"] or source_id,
                buyer_nickname=item["buyer_nickname"],
                title=f"与{item['buyer_nickname'] or source_id}的会话",
                created_at=_as_datetime(item["created_at"]) or utc_now(),
                updated_at=_as_datetime(item["created_at"]) or utc_now(),
            )
            db.add(record)
            db.flush()
            counts["conversations"] += 1
        conversations[source_id] = record

    for item in payload["orders"]:
        product_id = item["product_external_id"]
        if product_id:
            product = db.scalar(select(Product).where(Product.external_id == product_id))
            if product is None:
                db.add(
                    Product(
                        external_id=product_id,
                        sku=product_id,
                        name=item["product_name"] or product_id,
                        price=Decimal(item["unit_price"]) if item["unit_price"] else None,
                        extra={"source": "workbook"},
                    )
                )
                counts["products"] += 1
        values = {
            "customer_id": item["buyer_nickname"] or item["conversation_source_id"],
            "buyer_nickname": item["buyer_nickname"],
            "product_external_id": product_id,
            "conversation_id": conversations[item["conversation_source_id"]].id,
            "status": item["status"],
            "quantity": item["quantity"],
            "total_amount": item["total_amount"],
            "unit_price": item["unit_price"],
            "product_name": item["product_name"],
            "logistics_company": item["logistics_company"],
            "logistics_no": item["logistics_no"],
            "ordered_at": _as_datetime(item["ordered_at"]),
            "paid_at": _as_datetime(item["paid_at"]),
            "payment_stage": item["payment_stage"],
            "shipped_at": _as_datetime(item["shipped_at"]),
            "extra": item["extra"],
            "updated_at": utc_now(),
        }
        record = db.scalar(select(Order).where(Order.external_id == item["external_id"]))
        if record is None:
            db.add(Order(external_id=item["external_id"], **values))
            counts["orders"] += 1
        else:
            for key, value in values.items():
                setattr(record, key, value)
    db.flush()

    detail_models = {
        "replacement_exchange": ReplacementDetail,
        "offline_payment": OfflinePaymentDetail,
        "logistics": LogisticsDetail,
        "adverse_reaction": AdverseReactionDetail,
        "after_sale_return": ReturnDetail,
    }
    for item in payload["work_orders"]:
        values = {
            "ticket_type": item["ticket_type"],
            "conversation_id": conversations[item["conversation_source_id"]].id,
            "order_external_id": item["order_external_id"],
            "customer_id": item["buyer_nickname"] or item["conversation_source_id"],
            "buyer_nickname": item["buyer_nickname"],
            "status": item["status"],
            "source_status": item["source_status"],
            "assignee": item["assignee"],
            "description": item.get("description"),
            "opened_at": _as_datetime(item["opened_at"]),
            "closed_at": _as_datetime(item["closed_at"]),
            "source_extra": item["source_extra"],
            "updated_at": utc_now(),
        }
        record = db.scalar(select(WorkOrder).where(WorkOrder.external_id == item["external_id"]))
        if record is None:
            record = WorkOrder(external_id=item["external_id"], **values)
            db.add(record)
            db.flush()
            counts["work_orders"] += 1
        else:
            for key, value in values.items():
                setattr(record, key, value)
        _upsert_detail(db, detail_models[item["ticket_type"]], record.id, item["detail"])

    for item in payload["messages"]:
        values = {
            "conversation_id": conversations[item["conversation_source_id"]].id,
            "sender_role": item["sender_role"],
            "sequence_no": item["sequence_no"],
            "message_type": item["message_type"],
            "content": item["content"],
            "raw_content": item["raw_content"],
            "media_url": item["media_url"],
            "related_order_external_id": item["related_order_external_id"],
            "related_work_order_external_id": item["related_work_order_external_id"],
            "created_at": _as_datetime(item["created_at"]) or utc_now(),
        }
        record = db.scalar(
            select(Message).where(Message.source_external_id == item["source_external_id"])
        )
        if record is None:
            db.add(Message(source_external_id=item["source_external_id"], **values))
            counts["messages"] += 1
        else:
            for key, value in values.items():
                setattr(record, key, value)

    summary = dict(batch.summary)
    result = {"created": counts, "committed_at": utc_now().isoformat()}
    summary["commit_result"] = result
    batch.summary = summary
    batch.status = "committed"
    batch.committed_at = utc_now()
    db.commit()
    return result
